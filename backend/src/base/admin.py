import openpyxl
from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from django.utils.html import format_html

from base.models import City, Skill, Specialization, Position, Institution, SpecializationGroup, EmployeeCompany, \
    Country

try:
    from django_celery_results.admin import TaskResultAdmin

    is_task_result_defined = True
except Exception:
    is_task_result_defined = False

from django_celery_results.models import TaskResult

if is_task_result_defined:
    class CustomTaskResultAdmin(TaskResultAdmin):

        def task_input_and_output(self, obj):
            task_args = obj.task_args if obj.task_args else ""
            task_kwargs = obj.task_kwargs if obj.task_kwargs else ""
            result = obj.result if obj.result else ""
            decoded_args = task_args.encode().decode('unicode_escape')
            decoded_kwargs = task_kwargs.encode().decode('unicode_escape')
            decoded_result = result.encode().decode('unicode_escape')
            return format_html("<pre>Task Args: {}</pre><br><pre>Task Kwargs: {}</pre><br><pre>Task Result: {}</pre>",
                               decoded_args, decoded_kwargs, decoded_result)

        task_input_and_output.short_description = 'Task Input and Output'
        list_display = ('task_id', 'task_name', 'task_input_and_output', 'date_done', 'status')


    # Unregister the default admin and register the custom one
    admin.site.unregister(TaskResult)
    admin.site.register(TaskResult, CustomTaskResultAdmin)


class BaseAdmin(admin.ModelAdmin):
    def get_obj_id(self, obj):
        return f"✏️ {obj.id}"

    @admin.action(description='Обновить с вызовом сигнала')
    def call_update_signal(self, request, queryset):
        for query in queryset:
            query.updated_at = timezone.now()
            query.save()

    @admin.action(description='Экспорт в xlsx')
    def export_to_xlsx(self, request, queryset):
        wb = openpyxl.Workbook()
        default_sheet = wb.active
        ws = wb.create_sheet("Данные")
        wb.remove_sheet(default_sheet)
        file_name = f"{queryset.model.__name__}.xlsx"
        model_fields = queryset.model._meta.get_fields()
        for index, field in enumerate(model_fields):
            if hasattr(field, 'verbose_name') and field.verbose_name:
                field_name = field.verbose_name
            else:
                field_name = field.name
            ws.cell(row=1, column=index + 1).value = field_name
        for row_index, query in enumerate(queryset, start=2):
            for column_index, field in enumerate(model_fields, start=1):
                if field.is_relation:
                    try:
                        field_value = list(getattr(query, field.name).all().values_list('id', flat=True))
                    except AttributeError:
                        try:
                            field_value = list(getattr(query, f"{field.name}_set").all().values_list('id', flat=True))
                        except AttributeError:
                            field_value = None
                else:
                    field_value = getattr(query, field.name)

                if field_value:
                    field_value = str(field_value)
                else:
                    field_value = "-"
                ws.cell(row=row_index, column=column_index).value = field_value

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; '
                                       f'filename="{file_name}"'},
        )
        wb.save(response)

        return response

    readonly_fields = ['id', 'created_at', 'updated_at']
    actions = [call_update_signal, export_to_xlsx]


class VerificationAdmin(BaseAdmin):
    list_display = ('id', 'name', 'is_verified', 'created_at', 'updated_at')
    list_filter = ('is_verified',)
    search_fields = ('name',)
    readonly_fields = ('id', 'created_at', 'updated_at')
    fieldsets = (
        (None, {
            'fields': ('id', 'name', 'is_verified')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
        }),
    )


@admin.register(Country)
class CountryAdmin(VerificationAdmin):
    ordering = ['name']


@admin.register(City)
class CityAdmin(VerificationAdmin):
    ordering = ['name']


@admin.register(Skill)
class SkillAdmin(VerificationAdmin):
    ordering = ['name']


@admin.register(SpecializationGroup)
class SpecializationGroupAdmin(BaseAdmin):
    list_display = ('id', 'name', 'created_at', 'updated_at')
    search_fields = ('name',)
    ordering = ['name']
    fieldsets = (
        (None, {
            'fields': ('id', 'name')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
        }),
    )


@admin.register(Specialization)
class SpecializationAdmin(BaseAdmin):
    list_display = ('id', 'name', 'group', 'created_at', 'updated_at')
    search_fields = ('name', 'group__name')
    autocomplete_fields = ('group',)
    ordering = ['name']
    fieldsets = (
        (None, {
            'fields': ('id', 'name', 'group')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
        }),
    )


@admin.register(Position)
class PositionAdmin(VerificationAdmin):
    ordering = ['name']


@admin.register(Institution)
class InstitutionAdmin(VerificationAdmin):
    ordering = ['name']


@admin.register(EmployeeCompany)
class EmployeeCompanyAdmin(BaseAdmin):
    list_display = ('id', 'name', 'is_verified', 'created_at', 'updated_at')
    search_fields = ('name',)
    ordering = ['name']
    fieldsets = (
        (None, {
            'fields': ('id', 'name', 'is_verified')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
        }),
    )
